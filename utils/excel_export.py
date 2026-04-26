"""
Color-coded Excel export for classified job application emails.
Uses openpyxl to produce an .xlsx file where each row is colored
by its predicted category.
"""

from typing import Dict, Optional
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference


# Category -> (fill hex, font hex)
# Fills are light tints so black text is readable; font color darkens for contrast.
CATEGORY_COLORS: Dict[str, Dict[str, str]] = {
    "acceptance":      {"fill": "C6EFCE", "font": "006100"},  # green
    "rejection":       {"fill": "FFC7CE", "font": "9C0006"},  # red
    "interview":       {"fill": "BDD7EE", "font": "1F4E79"},  # blue
    "action_required": {"fill": "FCE4D6", "font": "974706"},  # orange
    "in_process":      {"fill": "FFF2CC", "font": "7F6000"},  # yellow
    "unrelated":       {"fill": "D9D9D9", "font": "404040"},  # gray
}

DEFAULT_COLOR = {"fill": "FFFFFF", "font": "000000"}

DISPLAY_NAMES: Dict[str, str] = {
    "action_required": "Action Required",
    "in_process": "In Progress",
    "acceptance": "Acceptance",
    "rejection": "Rejection",
    "interview": "Interview",
    "unrelated": "Unrelated",
}

# Columns to include in the Excel output (in order).
# Falls back gracefully if a column doesn't exist in the DataFrame.
EXPORT_COLUMNS = [
    "company",
    "subject",
    "date_only",
    "final_label",
    "rule_label",
    "rule_confidence",
    "ensemble_confidence",
    "ensemble_top2_gap",
    "review_status",
    "sentiment_label",
    "sentiment_compound",
    "extracted_role",
    "contact_person",
    "contact_email",
    "dates_mentioned",
    "summary",
    "email_body",
]

# Human-friendly header names
COLUMN_HEADERS = {
    "company": "Company",
    "subject": "Subject",
    "date_only": "Date",
    "final_label": "Category",
    "rule_label": "Rule Label",
    "rule_confidence": "Rule Confidence",
    "ensemble_confidence": "Model Confidence",
    "ensemble_top2_gap": "Top-2 Gap",
    "review_status": "Review Status",
    "sentiment_label": "Sentiment",
    "sentiment_compound": "Sentiment Score",
    "extracted_role": "Job Role",
    "contact_person": "Contact Person",
    "contact_email": "Contact Email",
    "dates_mentioned": "Dates Mentioned",
    "summary": "Summary",
    "email_body": "Email Body",
}


def _make_fill(hex_color: str) -> PatternFill:
    if len(hex_color) == 6:
        hex_color = f"FF{hex_color}"
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _make_font(hex_color: str, bold: bool = False, size: int = 11) -> Font:
    if len(hex_color) == 6:
        hex_color = f"FF{hex_color}"
    return Font(color=hex_color, bold=bold, size=size)


def export_to_excel(
    df: pd.DataFrame,
    output_path: str,
    label_column: str = "final_label",
    sheet_name: str = "Classified Emails",
) -> str:
    """
    Export the classified email DataFrame to a color-coded Excel file.

    Each row is tinted according to its category in ``label_column``.
    A legend sheet is appended showing the color key.

    Parameters
    ----------
    df : pd.DataFrame
        The enriched DataFrame from the pipeline (must contain ``label_column``).
    output_path : str
        Destination .xlsx path.
    label_column : str
        Column that holds the predicted category.
    sheet_name : str
        Name for the main data sheet.

    Returns
    -------
    str
        The absolute path of the written file.
    """
    if label_column not in df.columns:
        raise ValueError(f"Label column '{label_column}' not found in DataFrame. "
                         f"Available columns: {list(df.columns)}")

    # Select and order columns that exist
    cols = [c for c in EXPORT_COLUMNS if c in df.columns]
    if label_column not in cols:
        cols.insert(0, label_column)
    export_df = df[cols].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── Header row ──
    thin_border = Border(
        bottom=Side(style="thin", color="000000"),
    )
    header_fill = _make_fill("4472C4")
    header_font = _make_font("FFFFFF", bold=True, size=12)

    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx,
                       value=COLUMN_HEADERS.get(col_name, col_name))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = thin_border

    # ── Data rows ──
    label_col_idx = cols.index(label_column)

    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=2):
        label = str(row.get(label_column, ""))
        colors = CATEGORY_COLORS.get(label, DEFAULT_COLOR)
        row_fill = _make_fill(colors["fill"])
        row_font = _make_font(colors["font"])

        for col_idx, col_name in enumerate(cols, start=1):
            value = row[col_name]
            # Truncate very long text for readability
            if isinstance(value, str) and len(value) > 32767:
                value = value[:32767]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill
            cell.font = row_font
            cell.alignment = Alignment(vertical="top", wrap_text=(col_name in ("summary", "email_body")))

    # ── Auto-fit column widths (approximate) ──
    for col_idx, col_name in enumerate(cols, start=1):
        if col_name == "email_body":
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 60
        elif col_name == "summary":
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 50
        elif col_name == "subject":
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 40
        elif col_name in ("company", "extracted_role", "contact_person"):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20
        elif col_name in ("final_label", "rule_label", "sentiment_label", "review_status"):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18
        else:
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15

    # Freeze top row
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = ws.dimensions

    # ── Legend sheet ──
    _add_legend_sheet(wb)

    # ── Summary sheet ──
    _add_summary_sheet(wb, df, label_column)

    # ── Chart sheets ──
    _add_category_chart_sheet(wb, df, label_column)
    _add_sentiment_chart_sheet(wb, df)

    # Write
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return os.path.abspath(output_path)


def _add_legend_sheet(wb: Workbook):
    """Append a 'Legend' sheet explaining the color coding."""
    ws = wb.create_sheet(title="Legend")

    ws.cell(row=1, column=1, value="Category Color Legend").font = Font(bold=True, size=14)
    ws.merge_cells("A1:C1")

    headers = ["Category", "Description", "Color Sample"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=11)

    descriptions = {
        "acceptance": "Job offer extended to the candidate",
        "rejection": "Application declined by the company",
        "interview": "Interview invitation or scheduling",
        "action_required": "Assessment, form, or task to complete",
        "in_process": "Application received, under review",
        "unrelated": "Newsletter, job alert, or portal email",
    }

    for row_idx, (cat, colors) in enumerate(CATEGORY_COLORS.items(), start=4):
        display = DISPLAY_NAMES.get(cat, cat)
        ws.cell(row=row_idx, column=1, value=display).font = _make_font(colors["font"], bold=True)
        ws.cell(row=row_idx, column=2, value=descriptions.get(cat, ""))

        sample = ws.cell(row=row_idx, column=3, value=f"  {display}  ")
        sample.fill = _make_fill(colors["fill"])
        sample.font = _make_font(colors["font"])

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20


def _add_summary_sheet(wb: Workbook, df: pd.DataFrame, label_column: str):
    """Append a 'Summary' sheet with category counts, percentages,
    and (when available) a sentiment cross-tab broken down per category."""
    ws = wb.create_sheet(title="Summary")

    # ── Title ──
    title_cell = ws.cell(row=1, column=1, value="Classification Summary")
    title_cell.font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws.cell(row=2, column=1, value=f"Total emails: {len(df)}").font = Font(size=11)

    # ── Category counts table ──
    header_fill = _make_fill("4472C4")
    header_font_white = _make_font("FFFFFF", bold=True)

    for col, h in enumerate(["Category", "Count", "Percentage"], start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font_white

    counts = df[label_column].value_counts()
    total = len(df)

    last_counts_row = 4
    for row_idx, (cat, count) in enumerate(counts.items(), start=5):
        colors = CATEGORY_COLORS.get(cat, DEFAULT_COLOR)
        row_fill = _make_fill(colors["fill"])

        cell_cat = ws.cell(row=row_idx, column=1, value=DISPLAY_NAMES.get(cat, cat))
        cell_cat.fill = row_fill
        cell_cat.font = _make_font(colors["font"], bold=True)

        cell_count = ws.cell(row=row_idx, column=2, value=count)
        cell_count.fill = row_fill
        cell_count.font = _make_font(colors["font"])

        cell_pct = ws.cell(row=row_idx, column=3, value=f"{count / total * 100:.1f}%")
        cell_pct.fill = row_fill
        cell_pct.font = _make_font(colors["font"])

        last_counts_row = row_idx

    # ── Sentiment cross-tab (optional — only when sentiment_label exists) ──
    if "sentiment_label" in df.columns and df["sentiment_label"].notna().any():
        start_row = last_counts_row + 3

        title_cell = ws.cell(row=start_row, column=1,
                             value="Sentiment Breakdown by Category")
        title_cell.font = Font(bold=True, size=13)
        ws.merge_cells(
            start_row=start_row, start_column=1,
            end_row=start_row, end_column=6,
        )

        # Build cross-tab: index=category, columns=sentiment_label
        xtab = (
            df.groupby([label_column, "sentiment_label"])
            .size()
            .unstack(fill_value=0)
        )

        # Canonical column order
        sentiment_order = [c for c in ["positive", "neutral", "negative"]
                           if c in xtab.columns]
        other_cols = [c for c in xtab.columns if c not in sentiment_order]
        xtab = xtab[sentiment_order + other_cols]

        # Sentiment header colors
        sentiment_colors = {
            "positive": ("C6EFCE", "006100"),
            "neutral":  ("FFF2CC", "7F6000"),
            "negative": ("FFC7CE", "9C0006"),
        }
        default_sent_color = ("D9D9D9", "404040")

        header_row = start_row + 2
        ws.cell(row=header_row, column=1, value="Category").font = _make_font("FFFFFF", bold=True)
        ws.cell(row=header_row, column=1).fill = header_fill

        for col_idx, sent_label in enumerate(xtab.columns, start=2):
            fill_hex, font_hex = sentiment_colors.get(sent_label, default_sent_color)
            cell = ws.cell(row=header_row, column=col_idx,
                           value=sent_label.capitalize())
            cell.fill = _make_fill(fill_hex)
            cell.font = _make_font(font_hex, bold=True)

        # Add "Total" header
        total_col = len(xtab.columns) + 2
        ws.cell(row=header_row, column=total_col, value="Total").fill = header_fill
        ws.cell(row=header_row, column=total_col).font = _make_font("FFFFFF", bold=True)

        for row_offset, (cat, row_data) in enumerate(xtab.iterrows(), start=1):
            data_row = header_row + row_offset
            colors = CATEGORY_COLORS.get(cat, DEFAULT_COLOR)
            cat_cell = ws.cell(row=data_row, column=1, value=DISPLAY_NAMES.get(cat, cat))
            cat_cell.fill = _make_fill(colors["fill"])
            cat_cell.font = _make_font(colors["font"], bold=True)

            row_total = int(row_data.sum())
            for col_idx, sent_label in enumerate(xtab.columns, start=2):
                n = int(row_data[sent_label])
                pct = n / row_total * 100 if row_total else 0
                cell = ws.cell(row=data_row, column=col_idx,
                               value=f"{n} ({pct:.0f}%)")
                cell.fill = _make_fill(colors["fill"])
                cell.font = _make_font(colors["font"])

            total_cell = ws.cell(row=data_row, column=total_col, value=row_total)
            total_cell.fill = _make_fill(colors["fill"])
            total_cell.font = _make_font(colors["font"], bold=True)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12


def _add_category_chart_sheet(wb: Workbook, df: pd.DataFrame, label_column: str):
    """Append a 'Category Chart' sheet with a pie chart of category distribution."""
    ws = wb.create_sheet(title="Category Chart")

    counts = df[label_column].value_counts()
    ws.cell(row=1, column=1, value="Category")
    ws.cell(row=1, column=2, value="Count")
    for i, (cat, count) in enumerate(counts.items(), start=2):
        ws.cell(row=i, column=1, value=DISPLAY_NAMES.get(cat, cat))
        ws.cell(row=i, column=2, value=int(count))

    n = len(counts)
    pie = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    data = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Category Distribution"
    pie.style = 10

    ws.add_chart(pie, "D2")
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10


def _add_sentiment_chart_sheet(wb: Workbook, df: pd.DataFrame):
    """Append a 'Sentiment Chart' sheet with a bar chart of sentiment distribution."""
    ws = wb.create_sheet(title="Sentiment Chart")

    if "sentiment_label" not in df.columns or df["sentiment_label"].isna().all():
        ws.cell(row=1, column=1, value="No sentiment data available")
        return

    counts = df["sentiment_label"].value_counts()
    sent_order = [s for s in ["positive", "neutral", "negative"] if s in counts.index]
    other = [s for s in counts.index if s not in sent_order]
    ordered = sent_order + other

    ws.cell(row=1, column=1, value="Sentiment")
    ws.cell(row=1, column=2, value="Count")
    for i, sent in enumerate(ordered, start=2):
        ws.cell(row=i, column=1, value=sent.capitalize())
        ws.cell(row=i, column=2, value=int(counts.get(sent, 0)))

    n = len(ordered)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Sentiment Overview"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Sentiment"
    bar.style = 10

    data = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)

    ws.add_chart(bar, "D2")
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
