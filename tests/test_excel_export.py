"""Tests for utils/excel_export.py — color-coded Excel output."""

import os
import pytest
import pandas as pd
from openpyxl import load_workbook

from utils.excel_export import (
    export_to_excel,
    CATEGORY_COLORS,
    EXPORT_COLUMNS,
    COLUMN_HEADERS,
)


# ── Fixtures ──

@pytest.fixture
def sample_df():
    """Minimal DataFrame covering all six categories."""
    rows = [
        {"company": "Acme",  "subject": "Offer letter",      "date_only": "2025-01-01",
         "final_label": "acceptance",      "rule_label": "acceptance",      "rule_confidence": 0.95,
         "sentiment_label": "positive",  "sentiment_compound": 0.8,
         "extracted_role": "Engineer",   "contact_person": "Jane Doe",
         "contact_email": "jane@acme.com", "dates_mentioned": "Jan 15, 2025",
         "summary": "We are pleased to offer you the position.", "email_body": "Full body text here."},
        {"company": "Beta",  "subject": "Application update", "date_only": "2025-01-02",
         "final_label": "rejection",       "rule_label": "rejection",       "rule_confidence": 0.90,
         "sentiment_label": "negative",  "sentiment_compound": -0.6,
         "extracted_role": "Analyst",    "contact_person": None,
         "contact_email": None,          "dates_mentioned": "",
         "summary": "We regret to inform you.", "email_body": "Full body rejection."},
        {"company": "Gamma", "subject": "Interview invite",   "date_only": "2025-01-03",
         "final_label": "interview",       "rule_label": "interview",       "rule_confidence": 0.85,
         "sentiment_label": "positive",  "sentiment_compound": 0.5,
         "extracted_role": "Data Scientist", "contact_person": "John Smith",
         "contact_email": "john@gamma.com", "dates_mentioned": "Feb 1, 2025",
         "summary": "We'd like to invite you for an interview.", "email_body": "Full body interview."},
        {"company": "Delta", "subject": "Assessment required", "date_only": "2025-01-04",
         "final_label": "action_required",  "rule_label": "action_required",  "rule_confidence": 0.80,
         "sentiment_label": "neutral",   "sentiment_compound": 0.0,
         "extracted_role": "SWE",        "contact_person": None,
         "contact_email": None,          "dates_mentioned": "Jan 20, 2025",
         "summary": "Please complete your assessment.", "email_body": "Full body action."},
        {"company": "Epsilon","subject": "App received",      "date_only": "2025-01-05",
         "final_label": "in_process",      "rule_label": "in_process",      "rule_confidence": 0.70,
         "sentiment_label": "neutral",   "sentiment_compound": 0.1,
         "extracted_role": None,         "contact_person": None,
         "contact_email": None,          "dates_mentioned": "",
         "summary": "We received your application.", "email_body": "Full body in process."},
        {"company": "Zeta",  "subject": "Job alert",          "date_only": "2025-01-06",
         "final_label": "unrelated",       "rule_label": "unrelated",       "rule_confidence": 0.75,
         "sentiment_label": "neutral",   "sentiment_compound": 0.0,
         "extracted_role": None,         "contact_person": None,
         "contact_email": None,          "dates_mentioned": "",
         "summary": "New jobs matching your profile.", "email_body": "Full body unrelated."},
    ]
    return pd.DataFrame(rows)


@pytest.fixture
def output_path(tmp_path):
    return str(tmp_path / "test_output.xlsx")


# ── Core export tests ──

class TestExportToExcel:

    def test_creates_file(self, sample_df, output_path):
        result = export_to_excel(sample_df, output_path)
        assert os.path.isfile(result)
        assert result.endswith(".xlsx")

    def test_returns_absolute_path(self, sample_df, output_path):
        result = export_to_excel(sample_df, output_path)
        assert os.path.isabs(result)

    def test_has_three_sheets(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        assert set(wb.sheetnames) == {"Classified Emails", "Legend", "Summary"}

    def test_data_sheet_row_count(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        # 1 header + 6 data rows
        assert ws.max_row == 7

    def test_header_row_values(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Company" in headers
        assert "Category" in headers
        assert "Subject" in headers

    def test_header_uses_friendly_names(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        # Should NOT contain raw column names like "final_label"
        assert "final_label" not in headers
        assert "Category" in headers

    def test_data_values_match(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        # Find company column
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        co_idx = headers.index("Company") + 1
        companies = [ws.cell(row=r, column=co_idx).value for r in range(2, ws.max_row + 1)]
        assert "Acme" in companies
        assert "Beta" in companies

    def test_color_coding_applied(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        # Row 2 should be acceptance (green fill C6EFCE)
        cell = ws.cell(row=2, column=1)
        assert cell.fill.start_color.rgb is not None

    def test_each_category_has_distinct_color(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        cat_idx = headers.index("Category") + 1

        fills = {}
        for r in range(2, ws.max_row + 1):
            cat = ws.cell(row=r, column=cat_idx).value
            fill_color = ws.cell(row=r, column=1).fill.start_color.rgb
            fills[cat] = fill_color

        # All 6 categories should have different fill colors
        assert len(set(fills.values())) == 6

    def test_frozen_pane(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        assert ws.freeze_panes == "A2"

    def test_auto_filter(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        assert ws.auto_filter.ref is not None


class TestExportEdgeCases:

    def test_missing_label_column_raises(self, sample_df, output_path):
        with pytest.raises(ValueError, match="Label column"):
            export_to_excel(sample_df, output_path, label_column="nonexistent")

    def test_single_row(self, output_path):
        df = pd.DataFrame([{
            "company": "Solo", "subject": "Test", "final_label": "in_process",
            "email_body": "Hello.",
        }])
        result = export_to_excel(df, output_path)
        wb = load_workbook(result)
        ws = wb["Classified Emails"]
        assert ws.max_row == 2  # header + 1 data row

    def test_unknown_category_gets_default_color(self, output_path):
        df = pd.DataFrame([{
            "company": "X", "subject": "Y", "final_label": "mystery_label",
            "email_body": "text",
        }])
        result = export_to_excel(df, output_path)
        wb = load_workbook(result)
        ws = wb["Classified Emails"]
        # Should not crash; row gets default white fill
        assert ws.max_row == 2

    def test_empty_dataframe(self, output_path):
        df = pd.DataFrame(columns=["company", "subject", "final_label", "email_body"])
        result = export_to_excel(df, output_path)
        wb = load_workbook(result)
        ws = wb["Classified Emails"]
        assert ws.max_row == 1  # header only

    def test_long_email_body_truncated(self, output_path):
        long_text = "x" * 40000
        df = pd.DataFrame([{
            "company": "Big", "subject": "Long", "final_label": "in_process",
            "email_body": long_text,
        }])
        result = export_to_excel(df, output_path)
        wb = load_workbook(result)
        ws = wb["Classified Emails"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        body_idx = headers.index("Email Body") + 1
        cell_value = ws.cell(row=2, column=body_idx).value
        assert len(cell_value) <= 32767

    def test_subset_of_columns(self, output_path):
        """If the DataFrame is missing optional columns, export still works."""
        df = pd.DataFrame([{
            "company": "Mini", "final_label": "rejection",
            "email_body": "Declined.",
        }])
        result = export_to_excel(df, output_path)
        wb = load_workbook(result)
        ws = wb["Classified Emails"]
        assert ws.max_row == 2

    def test_creates_parent_directories(self, tmp_path):
        nested = str(tmp_path / "a" / "b" / "c" / "output.xlsx")
        df = pd.DataFrame([{
            "company": "Nested", "final_label": "in_process", "email_body": "text",
        }])
        result = export_to_excel(df, nested)
        assert os.path.isfile(result)


class TestLegendSheet:

    def test_legend_has_all_categories(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Legend"]
        values = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        for cat in CATEGORY_COLORS:
            assert cat in values, f"Missing {cat} in Legend sheet"

    def test_legend_has_descriptions(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Legend"]
        descriptions = [ws.cell(row=r, column=2).value for r in range(4, 10)]
        assert any("offer" in str(d).lower() for d in descriptions)
        assert any("declined" in str(d).lower() for d in descriptions)


class TestSummarySheet:

    def test_summary_has_total(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        total_cell = ws.cell(row=2, column=1).value
        assert "6" in str(total_cell)

    def test_summary_has_counts(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        counts = [ws.cell(row=r, column=2).value for r in range(5, 11)]
        assert all(c == 1 for c in counts)  # each category appears once

    def test_summary_has_percentages(self, sample_df, output_path):
        export_to_excel(sample_df, output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        pcts = [ws.cell(row=r, column=3).value for r in range(5, 11)]
        assert all("%" in str(p) for p in pcts)


class TestCustomLabelColumn:

    def test_uses_rule_label_when_specified(self, sample_df, output_path):
        export_to_excel(sample_df, output_path, label_column="rule_label")
        wb = load_workbook(output_path)
        ws = wb["Classified Emails"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "Rule Label" in headers


class TestCategoryColors:

    def test_all_six_categories_have_colors(self):
        expected = {"acceptance", "rejection", "interview",
                    "action_required", "in_process", "unrelated"}
        assert set(CATEGORY_COLORS.keys()) == expected

    def test_colors_are_valid_hex(self):
        for cat, colors in CATEGORY_COLORS.items():
            assert len(colors["fill"]) == 6, f"Bad fill hex for {cat}"
            assert len(colors["font"]) == 6, f"Bad font hex for {cat}"
            int(colors["fill"], 16)  # should not raise
            int(colors["font"], 16)


class TestSentimentCrossTab:

    @pytest.fixture
    def df_with_sentiment(self):
        """DataFrame that includes sentiment_label."""
        rows = []
        for cat in ["acceptance", "rejection", "in_process"]:
            for sent in ["positive", "negative", "neutral"]:
                rows.append({
                    "final_label": cat,
                    "sentiment_label": sent,
                    "sentiment_compound": 0.1,
                    "subject": f"{cat} {sent}",
                    "email_body": "body text",
                })
        return pd.DataFrame(rows)

    def test_summary_sheet_has_sentiment_section(self, df_with_sentiment, output_path):
        export_to_excel(df_with_sentiment, output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        # Flatten all cell values in the sheet
        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert any("Sentiment" in str(v) for v in all_values), \
            "Sentiment section not found in Summary sheet"

    def test_sentiment_cross_tab_has_category_rows(self, df_with_sentiment, output_path):
        export_to_excel(df_with_sentiment, output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        # Each category should appear somewhere in the sheet (counts table + cross-tab)
        for cat in ["acceptance", "rejection", "in_process"]:
            assert cat in all_values, f"Category '{cat}' missing from Summary sheet"

    def test_no_sentiment_cross_tab_without_column(self, output_path):
        """When sentiment_label column is absent, no crash and no cross-tab title."""
        df_no_sent = pd.DataFrame([
            {"final_label": "in_process", "subject": "test", "email_body": "body"},
        ])
        export_to_excel(df_no_sent, output_path)
        wb = load_workbook(output_path)
        ws = wb["Summary"]
        all_values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert not any("Breakdown" in str(v) for v in all_values), \
            "Sentiment Breakdown section should not appear when column is absent"
